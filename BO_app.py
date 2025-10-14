# ============================================================
# STREAMLIT BACK ORDER MANAGEMENT APP - OPTIMIZED
# Save this as: app.py
# Run with: streamlit run app.py
# ============================================================

import streamlit as st
import pandas as pd
import numpy as np
from datetime import datetime
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ============================================================
# PAGE CONFIG - MUST BE FIRST
# ============================================================

st.set_page_config(
    page_title="Back Order Management",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ============================================================
# CACHING FOR PERFORMANCE
# ============================================================

@st.cache_data
def load_and_clean_data(df):
    """Cache data cleaning operations"""
    df = df.copy()
    df['QOH'] = pd.to_numeric(df['QOH'], errors='coerce')
    df['Outstanding Amount'] = pd.to_numeric(df['Outstanding Amount'], errors='coerce')
    
    # Clean customer names
    df['Sell-to Customer Name'] = df['Sell-to Customer Name'].astype(str).str.strip()
    df = df[df['Sell-to Customer Name'] != '']
    df = df[df['Sell-to Customer Name'] != 'nan']
    df = df[df['Sell-to Customer Name'] != 'NaN']
    df = df[~df['Sell-to Customer Name'].str.isdigit()]
    
    # Remove rows with missing critical data
    df_clean = df.dropna(subset=['QOH', 'Sell-to Customer Name', 
                                   'Outstanding Amount', 'Mfg. Lead Name'])
    
    return df_clean

# ============================================================
# CUSTOM STYLING
# ============================================================

st.markdown("""
    <style>
    .header-title {
        color: #1f77b4;
        font-size: 28px;
        font-weight: bold;
        margin-bottom: 10px;
    }
    .back-order-item {
        background-color: #fff3cd;
        padding: 12px;
        margin: 8px 0;
        border-radius: 5px;
        border-left: 4px solid #ff9800;
    }
    .in-stock-item {
        background-color: #d4edda;
        padding: 12px;
        margin: 8px 0;
        border-radius: 5px;
        border-left: 4px solid #28a745;
    }
    .metric-small {
        background-color: #f0f2f6;
        padding: 10px;
        border-radius: 5px;
        margin: 5px 0;
    }
    </style>
""", unsafe_allow_html=True)

# ============================================================
# INITIALIZE SESSION STATE
# ============================================================

if 'df' not in st.session_state:
    st.session_state.df = None

if 'order_data_dict' not in st.session_state:
    st.session_state.order_data_dict = {}

# ============================================================
# HEADER
# ============================================================

st.markdown("<div class='header-title'>📦 Back Order Management Dashboard</div>", 
            unsafe_allow_html=True)
st.markdown("*Organized by customer with detailed order breakdown*")

st.markdown("---")

# ============================================================
# SIDEBAR - FILE UPLOAD
# ============================================================

with st.sidebar:
    st.header("📁 Upload Data")
    uploaded_file = st.file_uploader(
        "Upload Excel or CSV file",
        type=['xlsx', 'csv', 'xls'],
        help="Upload your ELS Global Back Order file"
    )
    
    if uploaded_file is not None:
        try:
            if uploaded_file.name.endswith('.xlsx') or uploaded_file.name.endswith('.xls'):
                df = pd.read_excel(uploaded_file)
            else:
                df = pd.read_csv(uploaded_file)
            
            st.session_state.df = df
            st.success("✅ File loaded successfully!")
            st.info(f"📊 {len(df)} rows | 🏷️ {len(df.columns)} columns")
            
        except Exception as e:
            st.error(f"❌ Error loading file: {str(e)}")
    
    st.divider()
    st.sidebar.header("ℹ️ About")
    st.sidebar.info(
        """
        This app helps you manage back orders by:
        - Organizing orders by customer
        - Showing stock status (in-stock vs back order)
        - Tracking dollar values
        - Recording reasons and comments per order
        - Exporting all data to Excel
        """
    )

# ============================================================
# MAIN CONTENT
# ============================================================

if st.session_state.df is not None:
    df = st.session_state.df.copy()
    
    try:
        # ============================================================
        # DATA PREPARATION - CACHED
        # ============================================================
        
        df_clean = load_and_clean_data(df)
        
        # Separate back orders and in-stock
        backorders = df_clean[df_clean['QOH'] == 0]
        instock = df_clean[df_clean['QOH'] > 0]
        
        # ============================================================
        # SUMMARY METRICS
        # ============================================================
        
        st.subheader("📊 Summary Metrics")
        
        col1, col2, col3, col4 = st.columns(4)
        
        with col1:
            st.metric(
                "Back Order Items",
                len(backorders),
                f"{(len(backorders)/len(df_clean)*100):.1f}% of all",
                border=True
            )
        
        with col2:
            back_order_value = backorders['Outstanding Amount'].sum()
            st.metric(
                "Back Order Value",
                f"${back_order_value:,.2f}",
                border=True
            )
        
        with col3:
            instock_value = instock['Outstanding Amount'].sum()
            st.metric(
                "In-Stock Value",
                f"${instock_value:,.2f}",
                border=True
            )
        
        with col4:
            total_value = df_clean['Outstanding Amount'].sum()
            st.metric(
                "Total Outstanding",
                f"${total_value:,.2f}",
                border=True
            )
        
        st.markdown("---")
        
        # ============================================================
        # FILTER OPTIONS
        # ============================================================
        
        st.subheader("🔍 Filter Options")
        
        col1, col2, col3, col4 = st.columns(4)
        
        with col1:
            selected_mfg = st.multiselect(
                "Manufacturing Lead",
                options=sorted(df_clean['Mfg. Lead Name'].dropna().unique()),
                help="Select one or more manufacturing leads"
            )
        
        with col2:
            min_value = st.number_input(
                "Min Item Dollar Value",
                min_value=0,
                value=0,
                step=100
            )
        
        with col3:
            max_value = st.number_input(
                "Max Item Dollar Value",
                min_value=0,
                value=int(df_clean['Outstanding Amount'].max()) + 1000,
                step=1000
            )
        
        with col4:
            stock_filter = st.selectbox(
                "Stock Status",
                ["All", "Back Order Only", "In Stock Only"],
                help="Filter by stock status"
            )
        
        # Apply filters
        filtered_df = df_clean.copy()
        
        if selected_mfg:
            filtered_df = filtered_df[filtered_df['Mfg. Lead Name'].isin(selected_mfg)]
        
        filtered_df = filtered_df[
            (filtered_df['Outstanding Amount'] >= min_value) &
            (filtered_df['Outstanding Amount'] <= max_value)
        ]
        
        if stock_filter == "Back Order Only":
            filtered_df = filtered_df[filtered_df['QOH'] == 0].copy()
        elif stock_filter == "In Stock Only":
            filtered_df = filtered_df[filtered_df['QOH'] > 0].copy()
        
        st.markdown("---")
        
        # ============================================================
        # CUSTOMERS LIST WITH DOLLAR FILTER
        # ============================================================
        
        # Calculate customer-level metrics
        customer_summary_list = []
        for cust in filtered_df['Sell-to Customer Name'].unique():
            cust_df = filtered_df[filtered_df['Sell-to Customer Name'] == cust]
            cust_total = cust_df['Outstanding Amount'].sum()
            customer_summary_list.append({
                'customer': cust,
                'total': cust_total
            })
        
        customer_summary_list = sorted(customer_summary_list, 
                                       key=lambda x: x['total'], reverse=True)
        
        # Customer-level dollar filter
        st.subheader("🏢 Filter Customers by Dollar Value")
        
        col1, col2 = st.columns(2)
        
        with col1:
            customer_min_value = st.number_input(
                "Customer Min. Total Outstanding",
                min_value=0,
                value=0,
                step=500,
                key="customer_min_filter"
            )
        
        with col2:
            max_cust_total = max([c['total'] for c in customer_summary_list]) if customer_summary_list else 0
            customer_max_value = st.number_input(
                "Customer Max. Total Outstanding",
                min_value=0,
                value=int(max_cust_total) + 1000,
                step=500,
                key="customer_max_filter"
            )
        
        # Apply customer dollar filter
        customer_list = [
            c['customer'] for c in customer_summary_list
            if customer_min_value <= c['total'] <= customer_max_value
        ]
        
        st.markdown("---")
        
        # ============================================================
        # CUSTOMERS SECTIONS
        # ============================================================
        
        if len(customer_list) > 0:
            st.subheader(f"👥 Customers ({len(customer_list)})")
            
            # Create expandable sections for each customer
            for customer_name in customer_list:
                # Get customer data
                customer_df = filtered_df[filtered_df['Sell-to Customer Name'] == customer_name].copy()
                
                # Calculate metrics
                total_customer_value = customer_df['Outstanding Amount'].sum()
                num_items = len(customer_df)
                back_order_items = len(customer_df[customer_df['QOH'] == 0])
                instock_items = len(customer_df[customer_df['QOH'] > 0])
                back_order_val = customer_df[customer_df['QOH'] == 0]['Outstanding Amount'].sum()
                instock_val = customer_df[customer_df['QOH'] > 0]['Outstanding Amount'].sum()
                mfg_leads_list = ', '.join(customer_df['Mfg. Lead Name'].unique())
                
                # Customer header with metrics
                with st.expander(
                    f"💼 {customer_name} | ${total_customer_value:,.2f} | {num_items} items",
                    expanded=False
                ):
                    
                    # ============================================================
                    # CUSTOMER OVERVIEW
                    # ============================================================
                    
                    st.write("**Customer Overview**")
                    col1, col2, col3, col4 = st.columns(4)
                    
                    with col1:
                        st.metric("Total Outstanding", f"${total_customer_value:,.2f}")
                    with col2:
                        st.metric("Total Items", num_items)
                    with col3:
                        st.metric("Back Order Items", back_order_items)
                    with col4:
                        st.metric("In-Stock Items", instock_items)
                    
                    st.write(f"**Manufacturing Leads:** {mfg_leads_list}")
                    
                    st.divider()
                    
                    # ============================================================
                    # DOLLAR VALUE SPLIT
                    # ============================================================
                    
                    st.write("**💰 Dollar Value Split**")
                    
                    col1, col2 = st.columns(2)
                    
                    with col1:
                        pct_bo = (back_order_val/total_customer_value*100) if total_customer_value > 0 else 0
                        st.markdown(f"""
                        <div class='metric-small'>
                            <b>Back Order Value:</b> ${back_order_val:,.2f} ({pct_bo:.1f}%)
                        </div>
                        """, unsafe_allow_html=True)
                    
                    with col2:
                        pct_is = (instock_val/total_customer_value*100) if total_customer_value > 0 else 0
                        st.markdown(f"""
                        <div class='metric-small'>
                            <b>In-Stock Value:</b> ${instock_val:,.2f} ({pct_is:.1f}%)
                        </div>
                        """, unsafe_allow_html=True)
                    
                    st.divider()
                    
                    # ============================================================
                    # IN-STOCK ITEMS
                    # ============================================================
                    
                    instock_customer = customer_df[customer_df['QOH'] > 0]
                    
                    if len(instock_customer) > 0:
                        st.write(f"**🟢 In-Stock Items ({len(instock_customer)} items)**")
                        
                        for idx, (_, item) in enumerate(instock_customer.iterrows()):
                            st.markdown(f"""
                            <div class='in-stock-item'>
                                <b>Item {idx+1}: {item['Item No']}</b><br>
                                📝 Description: {item['Desc'][:80]}<br>
                                💰 Outstanding Amount: <b>${item['Outstanding Amount']:,.2f}</b><br>
                                📦 QOH (In Stock): <b>{int(item['QOH'])} units</b><br>
                                🏭 Mfg Lead: {item['Mfg. Lead Name']}<br>
                                📅 Requested Delivery: {item.get('Requested Delivery Date', 'N/A')}<br>
                                📋 Order #: {item.get('Sales Order No', 'N/A')}
                            </div>
                            """, unsafe_allow_html=True)
                    
                    st.divider()
                    
                    # ============================================================
                    # BACK ORDER ITEMS
                    # ============================================================
                    
                    backorder_customer = customer_df[customer_df['QOH'] == 0]
                    
                    if len(backorder_customer) > 0:
                        st.write(f"**🔴 Back Order Items ({len(backorder_customer)} items) - QOH = 0**")
                        
                        for idx, (_, item) in enumerate(backorder_customer.iterrows()):
                            st.markdown(f"""
                            <div class='back-order-item'>
                                <b>Item {idx+1}: {item['Item No']}</b><br>
                                📝 Description: {item['Desc'][:80]}<br>
                                💰 Outstanding Amount: <b>${item['Outstanding Amount']:,.2f}</b><br>
                                📦 QOH: <b>0 units (OUT OF STOCK)</b><br>
                                🏭 Mfg Lead: {item['Mfg. Lead Name']}<br>
                                📅 Requested Delivery: {item.get('Requested Delivery Date', 'N/A')}<br>
                                📋 Order #: {item.get('Sales Order No', 'N/A')}<br>
                                ⏳ Status: <b>AWAITING MANUFACTURING</b>
                            </div>
                            """, unsafe_allow_html=True)
                    
                    st.divider()
                    
                    # ============================================================
                    # REASON & COMMENTS - PER SALES ORDER
                    # ============================================================
                    
                    st.write("**📝 Reason for Not Shipping & Comments - Per Sales Order**")
                    st.write("*Add reason and any comments for each individual sales order*")
                    
                    # Predefined reasons
                    predefined_reasons = [
                        "Manufacturing in progress",
                        "Waiting for approval",
                        "Quality check pending",
                        "Supply chain delay",
                        "Customer clarification needed",
                        "Payment pending",
                        "Customization in progress",
                        "Regulatory approval pending",
                        "Other"
                    ]
                    
                    # Display all orders for this customer with input fields
                    all_orders = customer_df[['Sales Order No', 'Item No', 'Desc', 
                                               'Outstanding Amount', 'QOH']].drop_duplicates(subset=['Sales Order No', 'Item No'])
                    
                    for idx, (_, order) in enumerate(all_orders.iterrows()):
                        order_key = f"{order['Sales Order No']}_{order['Item No']}"
                        stock_status = "IN STOCK" if order['QOH'] > 0 else "BACK ORDER"
                        
                        st.write(f"**Order {idx+1}:** {order['Sales Order No']} - Item: {order['Item No']}")
                        st.write(f"Description: {order['Desc'][:60]} | Amount: ${order['Outstanding Amount']:,.2f} | Status: {stock_status}")
                        
                        # Reason dropdown
                        selected_reason = st.selectbox(
                            "Select Reason:",
                            predefined_reasons,
                            key=f"order_reason_{order_key}_{id(customer_df)}"
                        )
                        
                        # Custom reason if "Other" selected
                        if selected_reason == "Other":
                            final_reason = st.text_input(
                                "Specify reason:",
                                key=f"order_custom_reason_{order_key}_{id(customer_df)}",
                                placeholder="Enter specific reason..."
                            )
                        else:
                            final_reason = selected_reason
                        
                        # Comments section - BELOW each order
                        st.write("**📌 Add Comments:**")
                        user_comments = st.text_area(
                            "Any additional comments or notes:",
                            key=f"order_comments_{order_key}_{id(customer_df)}",
                            height=100,
                            placeholder="Enter any comments, notes, or updates about this order..."
                        )
                        
                        # Save button
                        if st.button(
                            f"💾 Save for {order['Sales Order No']}",
                            key=f"save_order_{order_key}_{id(customer_df)}"
                        ):
                            st.session_state.order_data_dict[order_key] = {
                                'sales_order': order['Sales Order No'],
                                'item_no': order['Item No'],
                                'description': order['Desc'],
                                'reason': final_reason if final_reason else "Not specified",
                                'comments': user_comments,
                                'timestamp': datetime.now(),
                                'customer': customer_name,
                                'amount': order['Outstanding Amount'],
                                'stock_status': stock_status
                            }
                            st.success(f"✅ Data saved for {order['Sales Order No']}")
                        
                        st.divider()
        else:
            st.info("No customers match the selected filters.")
        
        st.markdown("---")
        
        # ============================================================
        # EXPORT & SUMMARY
        # ============================================================
        
        st.subheader("📥 Export & Reports")
        
        col1, col2, col3 = st.columns(3)
        
        with col1:
            # Export detailed back orders
            export_df = filtered_df[[
                'Sales Order No',
                'Sell-to Customer Name',
                'Item No',
                'Desc',
                'Outstanding Amount',
                'QOH',
                'Requested Delivery Date',
                'Mfg. Lead Name'
            ]].copy()
            
            export_df.columns = [
                'Order #',
                'Customer',
                'Item #',
                'Description',
                'Outstanding $',
                'QOH',
                'Delivery Date',
                'Mfg Lead'
            ]
            
            csv_export = export_df.to_csv(index=False)
            st.download_button(
                label="📋 Download Detailed Report (CSV)",
                data=csv_export,
                file_name=f"back_orders_detail_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
                mime="text/csv"
            )
        
        with col2:
            # Export customer summary
            summary_data = []
            for cust in customer_list:
                cust_df = filtered_df[filtered_df['Sell-to Customer Name'] == cust]
                cust_bo = cust_df[cust_df['QOH'] == 0]
                cust_is = cust_df[cust_df['QOH'] > 0]
                
                summary_data.append({
                    'Customer': cust,
                    'Total Outstanding $': cust_df['Outstanding Amount'].sum(),
                    'Total Items': len(cust_df),
                    'Back Order Items': len(cust_bo),
                    'Back Order $': cust_bo['Outstanding Amount'].sum(),
                    'In-Stock Items': len(cust_is),
                    'In-Stock $': cust_is['Outstanding Amount'].sum(),
                    'Mfg Leads': ', '.join(cust_df['Mfg. Lead Name'].unique())
                })
            
            if summary_data:
                summary_df = pd.DataFrame(summary_data)
                csv_summary = summary_df.to_csv(index=False)
                st.download_button(
                    label="📊 Download Summary (CSV)",
                    data=csv_summary,
                    file_name=f"back_orders_summary_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
                    mime="text/csv"
                )
            else:
                st.info("No data to export")
        
        with col3:
            st.write("")  # Spacing
        
        st.markdown("---")
        
        # ============================================================
        # EXPORT ALL ORDER DATA TO EXCEL
        # ============================================================
        
        st.subheader("📊 Export All Order Data with Comments")
        
        if st.session_state.order_data_dict:
            
            if st.button("📥 Generate Excel Export"):
                # Create Excel workbook
                wb = Workbook()
                ws = wb.active
                ws.title = "Order Data"
                
                # Define styles
                header_fill = PatternFill(start_color="1F77B4", end_color="1F77B4", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF")
                border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                # Headers
                headers = ['Customer', 'Sales Order', 'Item No', 'Description', 'Amount', 
                          'Stock Status', 'Reason', 'Comments', 'Recorded Date', 'Recorded Time']
                
                for col_num, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_num)
                    cell.value = header
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = border
                
                # Add data rows
                for row_num, (order_key, data) in enumerate(st.session_state.order_data_dict.items(), 2):
                    timestamp = data['timestamp']
                    
                    ws.cell(row=row_num, column=1).value = data['customer']
                    ws.cell(row=row_num, column=2).value = data['sales_order']
                    ws.cell(row=row_num, column=3).value = data['item_no']
                    ws.cell(row=row_num, column=4).value = data['description']
                    ws.cell(row=row_num, column=5).value = data['amount']
                    ws.cell(row=row_num, column=6).value = data['stock_status']
                    ws.cell(row=row_num, column=7).value = data['reason']
                    ws.cell(row=row_num, column=8).value = data['comments']
                    ws.cell(row=row_num, column=9).value = timestamp.strftime("%Y-%m-%d")
                    ws.cell(row=row_num, column=10).value = timestamp.strftime("%H:%M:%S")
                    
                    # Apply borders
                    for col_num in range(1, len(headers) + 1):
                        ws.cell(row=row_num, column=col_num).border = border
                    
                    # Set column width for comments
                    ws.column_dimensions['H'].width = 40
                
                # Set column widths
                ws.column_dimensions['A'].width = 25
                ws.column_dimensions['B'].width = 15
                ws.column_dimensions['C'].width = 15
                ws.column_dimensions['D'].width = 35
                ws.column_dimensions['E'].width = 12
                ws.column_dimensions['F'].width = 12
                ws.column_dimensions['G'].width = 25
                ws.column_dimensions['I'].width = 12
                ws.column_dimensions['J'].width = 12
                
                # Save to bytes
                excel_file = io.BytesIO()
                wb.save(excel_file)
                excel_file.seek(0)
                
                # Download button
                st.download_button(
                    label="💾 Download as Excel",
                    data=excel_file,
                    file_name=f"back_order_data_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
                
                st.success("✅ Excel file ready for download!")
        else:
            st.info("💡 No order data recorded yet. Add reasons and comments above to export.")
        
        st.markdown("---")
        
        # ============================================================
        # RECORDED DATA SUMMARY
        # ============================================================
        
        if st.session_state.order_data_dict:
            st.subheader("✅ Recorded Data - By Sales Order")
            
            for order_key, data in st.session_state.order_data_dict.items():
                with st.expander(f"📦 {data['sales_order']} - {data['item_no']} | ${data['amount']:,.2f}"):
                    col1, col2 = st.columns(2)
                    
                    with col1:
                        st.write(f"**Customer:** {data['customer']}")
                        st.write(f"**Description:** {data['description'][:60]}")
                        st.write(f"**Amount:** ${data['amount']:,.2f}")
                        st.write(f"**Stock Status:** {data['stock_status']}")
                    
                    with col2:
                        st.write(f"**Reason:** {data['reason']}")
                        st.write(f"**Comments:**")
                        st.write(data['comments'] if data['comments'] else "No comments added")
                    
                    st.caption(f"🕐 Recorded: {data['timestamp'].strftime('%Y-%m-%d %H:%M:%S')}")
    
    except Exception as e:
        st.error(f"❌ Error processing data: {str(e)}")
        import traceback
        st.write(traceback.format_exc())
        st.write("Please ensure your file has the required columns:")
        st.write("- QOH")
        st.write("- Sell-to Customer Name")
        st.write("- Outstanding Amount")
        st.write("- Mfg. Lead Name")

else:
    st.info(
        """
        👈 **Getting Started:**
        
        1. **Upload your file** using the sidebar (Excel or CSV)
        2. **Review your back orders** - organized by customer
        3. **See the split** between in-stock and back order items with dollar values
        4. **Add reason and comments** for each sales order
        5. **Export all data** to Excel with one click
        
        **Required columns in your file:**
        - QOH (Quantity on Hand)
        - Sell-to Customer Name
        - Outstanding Amount
        - Mfg. Lead Name
        - Sales Order No
        - Item No
        - Desc (Description)
        - Requested Delivery Date
        """
    )
